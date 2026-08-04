import os
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from extractor_oc.parser import OrdenCompra


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def _style_header(ws, headers, start_row=1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data(ws, row_num, col_count, wrap_cols=None):
    if wrap_cols is None:
        wrap_cols = {2}
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center" if col > 1 else "left",
            vertical="center",
            wrap_text=col in wrap_cols,
        )


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


def exportar_orden(oc: OrdenCompra, filepath: str) -> str:
    wb = Workbook()

    ws = wb.active
    ws.title = "Orden de Compra"

    row = 1
    ws.cell(row=row, column=1, value="ORDEN DE COMPRA POR CATÁLOGO ELECTRÓNICO")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    info = [
        ("Orden de Compra:", oc.orden_compra),
        ("Fecha de Aceptación:", oc.fecha_aceptacion),
        ("Nombre Comercial:", oc.nombre_comercial),
        ("Razón Social:", oc.razon_social),
        ("RUC:", oc.ruc),
        ("Administrador:", oc.administrador),
    ]

    for label, value in info:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=3, value=value)
        ws.cell(row=row, column=3).font = CELL_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1

    if oc.objeto_contratacion:
        ws.cell(row=row, column=1, value="Objeto:")
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=3, value=oc.objeto_contratacion)
        ws.cell(row=row, column=3).font = CELL_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 2

    headers = ["CPC", "Descripción", "Unidad", "Cantidad", "V. Unitario", "Subtotal", "Partida Presup."]
    _style_header(ws, headers, start_row=row)
    row += 1

    for item in oc.items:
        ws.cell(row=row, column=1, value=item.cpc)
        ws.cell(row=row, column=2, value=item.descripcion)
        ws.cell(row=row, column=3, value=item.unidad)
        ws.cell(row=row, column=4, value=item.cantidad)
        ws.cell(row=row, column=5, value=item.v_unitario)

        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=item.subtotal)
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=item.partida_presupuestaria)

        _style_data(ws, row, 7, wrap_cols={2})
        ws.cell(row=row, column=6).font = BOLD_FONT
        row += 1

    total_row = row
    ws.cell(row=total_row, column=5, value="V. TOTAL")
    ws.cell(row=total_row, column=5).font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    ws.cell(row=total_row, column=5).fill = TOTAL_FILL
    ws.cell(row=total_row, column=5).border = THIN_BORDER
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=6, value=oc.v_total)
    ws.cell(row=total_row, column=6).font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    ws.cell(row=total_row, column=6).fill = TOTAL_FILL
    ws.cell(row=total_row, column=6).number_format = '#,##0.00'
    ws.cell(row=total_row, column=6).border = THIN_BORDER

    col_widths = {1: 14, 2: 45, 3: 10, 4: 12, 5: 14, 6: 14, 7: 30}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(filepath)
    return filepath


def exportar_multiples(ordenes: List[OrdenCompra], filepath: str) -> str:
    wb = Workbook()

    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    headers_resumen = [
        "Orden de Compra", "Fecha", "Nombre Comercial", "Razón Social", "RUC",
        "Descripción", "Partida Presup.", "Administrador", "V. Total",
    ]
    _style_header(ws_resumen, headers_resumen)
    for i, oc in enumerate(ordenes, 2):
        desc = oc.items[0].descripcion if oc.items else ""
        if len(oc.items) > 1:
            desc += f" (+{len(oc.items) - 1} más)"
        partidas = list(dict.fromkeys(it.partida_presupuestaria for it in oc.items if it.partida_presupuestaria))
        ws_resumen.cell(row=i, column=1, value=oc.orden_compra)
        ws_resumen.cell(row=i, column=2, value=oc.fecha_aceptacion)
        ws_resumen.cell(row=i, column=3, value=oc.nombre_comercial)
        ws_resumen.cell(row=i, column=4, value=oc.razon_social)
        ws_resumen.cell(row=i, column=5, value=oc.ruc)
        ws_resumen.cell(row=i, column=6, value=desc)
        ws_resumen.cell(row=i, column=7, value=", ".join(partidas) if partidas else "")
        ws_resumen.cell(row=i, column=8, value=oc.administrador)
        ws_resumen.cell(row=i, column=9, value=oc.v_total)
        ws_resumen.cell(row=i, column=9).number_format = '#,##0.00'
        _style_data(ws_resumen, i, len(headers_resumen), wrap_cols={6})
    _auto_width(ws_resumen)

    ws_detalle = wb.create_sheet("Detalle")
    headers_detalle = [
        "Orden de Compra", "CPC", "Descripción", "Unidad",
        "Cantidad", "V. Unitario", "Subtotal", "Partida Presup.",
    ]
    _style_header(ws_detalle, headers_detalle)
    row = 2
    for oc in ordenes:
        for item in oc.items:
            ws_detalle.cell(row=row, column=1, value=oc.orden_compra)
            ws_detalle.cell(row=row, column=2, value=item.cpc)
            ws_detalle.cell(row=row, column=3, value=item.descripcion)
            ws_detalle.cell(row=row, column=4, value=item.unidad)
            ws_detalle.cell(row=row, column=5, value=item.cantidad)
            ws_detalle.cell(row=row, column=5).number_format = '#,##0.00'
            ws_detalle.cell(row=row, column=6, value=item.v_unitario)
            ws_detalle.cell(row=row, column=6).number_format = '#,##0.00'
            ws_detalle.cell(row=row, column=7, value=item.subtotal)
            ws_detalle.cell(row=row, column=7).number_format = '#,##0.00'
            ws_detalle.cell(row=row, column=8, value=item.partida_presupuestaria)
            _style_data(ws_detalle, row, len(headers_detalle), wrap_cols={3})
            row += 1
    _auto_width(ws_detalle)

    wb.save(filepath)
    return filepath
