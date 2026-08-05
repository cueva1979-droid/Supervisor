import re, os, sys, datetime, math
from collections import defaultdict
from typing import Optional, Dict, List
from sqlalchemy.orm import Session
from sqlalchemy import func
from models import Item, Record, Provider
from services.security import sanitize_excel

def _normalize(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'[^\w\s]', '', text)
    return text

def _compute_stats(items: List[Dict]) -> Dict:
    if not items:
        return {
            "total_productos": 0,
            "total_compras": 0,
            "producto_mas_comprado": "---",
            "producto_precio_mas_alto": "---",
            "producto_precio_mas_bajo": "---",
            "valor_total_acumulado": 0,
        }
    mas_comprado = max(items, key=lambda x: x["total_adquirido"])
    mas_alto = max(items, key=lambda x: x["precio_max"])
    mas_bajo = min(items, key=lambda x: x["precio_min"])
    total_compras = sum(it["compras"] for it in items)
    valor_total = sum(it["total_adquirido"] * it["precio_promedio"] for it in items)
    return {
        "total_productos": len(items),
        "total_compras": total_compras,
        "producto_mas_comprado": mas_comprado["descripcion"],
        "producto_precio_mas_alto": mas_alto["descripcion"],
        "producto_precio_mas_bajo": mas_bajo["descripcion"],
        "valor_total_acumulado": round(valor_total, 2),
    }

def _compute_charts(aggregated: List[Dict], detail_rows: List[Dict]) -> Dict:
    top = sorted(aggregated, key=lambda x: x["total_adquirido"], reverse=True)[:10]
    chart_top_products = [{"label": it["descripcion"][:30], "value": it["total_adquirido"]} for it in top]

    prices_by_date = defaultdict(list)
    for r in detail_rows:
        d = r.get("fecha") or "Sin fecha"
        prices_by_date[d].append(r["precio_unitario"])
    price_evo = [{"label": d, "value": round(sum(v)/len(v), 2)} for d, v in sorted(prices_by_date.items())]

    purchases_by_prov = defaultdict(int)
    for r in detail_rows:
        purchases_by_prov[r["proveedor"]] += 1
    sorted_prov = sorted(purchases_by_prov.items(), key=lambda x: x[1], reverse=True)[:10]
    chart_provider_purchases = [{"label": p, "value": c} for p, c in sorted_prov]

    return {
        "chart_top_products": chart_top_products,
        "chart_price_evolution": price_evo,
        "chart_provider_purchases": chart_provider_purchases,
    }

def get_product_report(
    db: Session,
    search: Optional[str] = None,
    producto: Optional[str] = None,
    proveedor: Optional[str] = None,
    codigo_cpc: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    precio_min: Optional[float] = None,
    precio_max: Optional[float] = None,
    codigo_proceso: Optional[str] = None,
    page: int = 1,
    per_page: int = 20,
) -> Dict:
    rows = (
        db.query(
            Item.codigo_cpc,
            Item.descripcion,
            Item.cantidad,
            Item.precio_unitario,
            Item.subtotal,
            Record.proveedor,
            Record.ruc,
            Record.codigo_proceso,
            Record.fecha,
            Record.numero_orden,
            Provider.nombre.label("provider_nombre"),
            Provider.ruc.label("provider_ruc"),
        )
        .join(Record, Item.record_id == Record.id)
        .join(Provider, Record.provider_id == Provider.id)
        .all()
    )

    detail_rows = []
    for r in rows:
        detail_rows.append({
            "codigo_cpc": r.codigo_cpc or "",
            "descripcion": r.descripcion or "",
            "cantidad": r.cantidad or 0,
            "precio_unitario": r.precio_unitario or 0,
            "subtotal": r.subtotal or 0,
            "proveedor": r.provider_nombre or r.proveedor or "",
            "ruc": r.provider_ruc or r.ruc or "",
            "codigo_proceso": r.codigo_proceso or "",
            "fecha": r.fecha or "",
            "numero_orden": r.numero_orden or "",
        })

    aggregated_map = {}
    for r in detail_rows:
        key = r["codigo_cpc"] + "||" + _normalize(r["descripcion"])
        if key not in aggregated_map:
            aggregated_map[key] = {
                "codigo_cpc": r["codigo_cpc"],
                "descripcion": r["descripcion"],
                "precios": [],
                "compras": 0,
                "total_adquirido": 0,
                "proveedores": set(),
            }
        g = aggregated_map[key]
        g["precios"].append(r["precio_unitario"])
        g["compras"] += 1
        g["total_adquirido"] += r["cantidad"]
        g["proveedores"].add(r["proveedor"])

    aggregated = []
    for key, g in aggregated_map.items():
        prices = g["precios"]
        aggregated.append({
            "codigo_cpc": g["codigo_cpc"],
            "descripcion": g["descripcion"],
            "precio_min": min(prices),
            "precio_max": max(prices),
            "precio_promedio": round(sum(prices) / len(prices), 2),
            "compras": g["compras"],
            "total_adquirido": round(g["total_adquirido"], 2),
            "proveedores": ", ".join(sorted(g["proveedores"])),
        })

    if search:
        term = search.lower()
        aggregated = [it for it in aggregated if term in it["descripcion"].lower() or term in it["codigo_cpc"].lower()]
        detail_rows = [r for r in detail_rows if term in r["descripcion"].lower() or term in r["codigo_cpc"].lower()]
    if producto:
        term = producto.lower()
        aggregated = [it for it in aggregated if term in it["descripcion"].lower()]
    if proveedor:
        term = proveedor.lower()
        aggregated = [it for it in aggregated if any(term in p.lower() for p in it["proveedores"])]
    if codigo_cpc:
        aggregated = [it for it in aggregated if codigo_cpc.lower() in it["codigo_cpc"].lower()]
    if codigo_proceso:
        aggregated = [it for it in aggregated if any(codigo_proceso.lower() in r["codigo_proceso"].lower() for r in detail_rows if r["codigo_cpc"] + "||" + _normalize(r["descripcion"]) in [it2["codigo_cpc"] + "||" + _normalize(it2["descripcion"]) for it2 in aggregated])]
    if fecha_desde:
        detail_rows = [r for r in detail_rows if r["fecha"] >= fecha_desde]
    if fecha_hasta:
        detail_rows = [r for r in detail_rows if r["fecha"] <= fecha_hasta]
    if precio_min is not None:
        aggregated = [it for it in aggregated if it["precio_promedio"] >= precio_min]
    if precio_max is not None:
        aggregated = [it for it in aggregated if it["precio_promedio"] <= precio_max]

    total = len(aggregated)
    start = (page - 1) * per_page
    end = start + per_page
    page_items = aggregated[start:end]

    stats = _compute_stats(aggregated)
    charts = _compute_charts(aggregated, detail_rows)

    return {
        "items": page_items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page),
        "stats": stats,
        "charts": charts,
        "_detail": detail_rows,
    }

def generate_product_excel(
    db: Session,
    search: Optional[str] = None,
    producto: Optional[str] = None,
    proveedor: Optional[str] = None,
    codigo_cpc: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    precio_min: Optional[float] = None,
    precio_max: Optional[float] = None,
    codigo_proceso: Optional[str] = None,
) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    report = get_product_report(
        db, search, producto, proveedor, codigo_cpc,
        fecha_desde, fecha_hasta, precio_min, precio_max,
        codigo_proceso, page=1, per_page=999999
    )

    detail_rows = report.get("_detail", [])

    wb = Workbook()
    hfont = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    # ---------- Sheet 1: Resumen Productos ----------
    ws1 = wb.active
    ws1.title = "Resumen Productos"
    headers1 = ["Código CPC", "Producto", "Precio Mín", "Precio Máx",
                "Precio Prom", "Compras", "Total Adquirido", "Proveedores"]
    for ci, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = thin
    for ri, it in enumerate(report["items"], 2):
        vals = [it["codigo_cpc"], it["descripcion"], it["precio_min"],
                it["precio_max"], it["precio_promedio"], it["compras"],
                it["total_adquirido"], it["proveedores"]]
        for ci, v in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=ci, value=sanitize_excel(v))
            c.border = thin
            c.alignment = Alignment(vertical="center")
    w1 = [18, 45, 14, 14, 14, 12, 16, 40]
    for ci, w in enumerate(w1, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.auto_filter.ref = f"A1:H{len(report['items']) + 1}"
    ws1.cell(row=len(report["items"]) + 3, column=1,
             value=f"Reporte generado el: {now}").font = Font(italic=True, color="666666")

    # ---------- Sheet 2: Detalle Completo ----------
    ws2 = wb.create_sheet("Detalle Completo")
    headers2 = ["Código CPC", "Producto", "Cantidad", "Precio Unitario",
                "Subtotal", "Proveedor", "RUC", "Código Proceso",
                "Fecha", "N° Orden"]
    for ci, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = thin
    for ri, r in enumerate(detail_rows, 2):
        vals = [r["codigo_cpc"], r["descripcion"], r["cantidad"],
                r["precio_unitario"], r["subtotal"], r["proveedor"],
                r["ruc"], r["codigo_proceso"], r["fecha"], r["numero_orden"]]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=sanitize_excel(v))
            c.border = thin
            c.alignment = Alignment(vertical="center")
    w2 = [18, 45, 12, 16, 16, 30, 18, 18, 16, 24]
    for ci, w in enumerate(w2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.auto_filter.ref = f"A1:J{len(detail_rows) + 1}"
    ws2.cell(row=len(detail_rows) + 3, column=1,
             value=f"Reporte generado el: {now}").font = Font(italic=True, color="666666")

    # ---------- Sheet 3: Estadisticas ----------
    ws3 = wb.create_sheet("Estadisticas")
    st = report["stats"]
    stats_data = [
        ("Indicador", "Valor"),
        ("Total Productos Registrados", st["total_productos"]),
        ("Total Compras Realizadas", st["total_compras"]),
        ("Producto Mas Comprado", st["producto_mas_comprado"]),
        ("Producto con Precio Mas Alto", st["producto_precio_mas_alto"]),
        ("Producto con Precio Mas Bajo", st["producto_precio_mas_bajo"]),
        ("Valor Total Acumulado", st["valor_total_acumulado"]),
    ]
    for ri, (label, val) in enumerate(stats_data, 1):
        c1 = ws3.cell(row=ri, column=1, value=label)
        c2 = ws3.cell(row=ri, column=2, value=sanitize_excel(val))
        if ri == 1:
            c1.font = hfont; c1.fill = hfill; c1.alignment = halign; c1.border = thin
            c2.font = hfont; c2.fill = hfill; c2.alignment = halign; c2.border = thin
        else:
            c1.font = Font(bold=True); c1.border = thin
            c2.border = thin
        c1.alignment = Alignment(vertical="center")
        c2.alignment = Alignment(vertical="center")
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 50

    row_offset = len(stats_data) + 3
    ws3.cell(row=row_offset, column=1, value="Top 10 Productos Mas Comprados").font = Font(bold=True, size=12)
    ws3.cell(row=row_offset + 1, column=1, value="Producto").font = Font(bold=True)
    ws3.cell(row=row_offset + 1, column=2, value="Total Adquirido").font = Font(bold=True)
    for ci in range(1, 3):
        ws3.cell(row=row_offset + 1, column=ci).border = thin
    top = sorted(report["items"], key=lambda x: x["total_adquirido"], reverse=True)[:10]
    for ri, it in enumerate(top, row_offset + 2):
        ws3.cell(row=ri, column=1, value=sanitize_excel(it["descripcion"][:60])).border = thin
        ws3.cell(row=ri, column=2, value=it["total_adquirido"]).border = thin

    row_offset2 = row_offset + len(top) + 3
    ws3.cell(row=row_offset2, column=1, value=f"Reporte generado el: {now}").font = Font(italic=True, color="666666")

    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        base = os.path.dirname(base)
    out_dir = os.path.join(base, "exports")
    os.makedirs(out_dir, exist_ok=True)
    filepath = os.path.join(out_dir, "Reporte_Productos_Precios.xlsx")
    wb.save(filepath)
    return filepath
