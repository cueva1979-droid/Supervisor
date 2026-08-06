import os
import sys
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from models import Record, Item, Provider
from services.security import sanitize_excel

def get_export_dir():
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
    export_dir = os.path.join(base, "exports")
    os.makedirs(export_dir, exist_ok=True)
    return export_dir

EXPORT_DIR = get_export_dir()

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def style_header(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

def style_data(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

def generate_excel(db: Session) -> str:
    wb = Workbook()
    records = db.query(Record).order_by(Record.fecha_procesamiento.desc()).all()
    providers = db.query(Provider).order_by(Provider.nombre).all()

    ws1 = wb.active
    ws1.title = "Resumen General"
    headers1 = ["Total Documentos", "Total Proveedores", "Total Órdenes", "Monto Total Acumulado"]
    style_header(ws1, headers1)
    total_ordenes = len(records)
    total_monto = sum(r.monto_total or 0 for r in records)
    data1 = [len(records), len(providers), total_ordenes, round(total_monto, 2)]
    for col, val in enumerate(data1, 1):
        ws1.cell(row=2, column=col, value=val)
    style_data(ws1, 2, len(headers1))
    auto_width(ws1)

    ws2 = wb.create_sheet("Proveedores")
    headers2 = ["ID", "Nombre", "RUC", "Código Proceso", "Teléfono", "Observaciones", "Fecha Creación"]
    style_header(ws2, headers2)
    for row_num, prov in enumerate(providers, 2):
        ws2.cell(row=row_num, column=1, value=prov.id)
        ws2.cell(row=row_num, column=2, value=sanitize_excel(prov.nombre))
        ws2.cell(row=row_num, column=3, value=sanitize_excel(prov.ruc))
        ws2.cell(row=row_num, column=4, value=sanitize_excel(prov.codigo_proceso or ""))
        ws2.cell(row=row_num, column=5, value=sanitize_excel(prov.telefono or ""))
        ws2.cell(row=row_num, column=6, value=sanitize_excel(prov.observaciones or ""))
        ws2.cell(row=row_num, column=7, value=prov.fecha_creacion.strftime("%d/%m/%Y %H:%M") if prov.fecha_creacion else "")
        style_data(ws2, row_num, len(headers2))
    auto_width(ws2)

    ws3 = wb.create_sheet("Órdenes")
    headers3 = ["ID", "Archivo", "Tipo", "Proveedor", "RUC", "Código Proceso", "N° Orden", "Fecha", "Objeto", "Monto Total", "Moneda", "Estado", "Fecha Procesamiento"]
    style_header(ws3, headers3)
    for row_num, rec in enumerate(records, 2):
        ws3.cell(row=row_num, column=1, value=rec.id)
        ws3.cell(row=row_num, column=2, value=sanitize_excel(rec.filename))
        ws3.cell(row=row_num, column=3, value=sanitize_excel(rec.file_type))
        ws3.cell(row=row_num, column=4, value=sanitize_excel(rec.proveedor or ""))
        ws3.cell(row=row_num, column=5, value=sanitize_excel(rec.ruc or ""))
        ws3.cell(row=row_num, column=6, value=sanitize_excel(rec.codigo_proceso or ""))
        ws3.cell(row=row_num, column=7, value=sanitize_excel(rec.numero_orden or ""))
        ws3.cell(row=row_num, column=8, value=sanitize_excel(rec.fecha or ""))
        ws3.cell(row=row_num, column=9, value=sanitize_excel(rec.objeto_contratacion or ""))
        ws3.cell(row=row_num, column=10, value=rec.monto_total or 0)
        ws3.cell(row=row_num, column=11, value=sanitize_excel(rec.moneda))
        ws3.cell(row=row_num, column=12, value=sanitize_excel(rec.estado))
        ws3.cell(row=row_num, column=13, value=rec.fecha_procesamiento.strftime("%d/%m/%Y %H:%M") if rec.fecha_procesamiento else "")
        style_data(ws3, row_num, len(headers3))
    auto_width(ws3)

    ws4 = wb.create_sheet("Ítems")
    headers4 = ["ID", "ID Orden", "Código CPC", "Descripción", "Cantidad", "Unidad", "Precio Unitario", "Subtotal", "Requiere Revisión"]
    style_header(ws4, headers4)
    row_num = 2
    for rec in records:
        for item in rec.items:
            ws4.cell(row=row_num, column=1, value=item.id)
            ws4.cell(row=row_num, column=2, value=rec.id)
            ws4.cell(row=row_num, column=3, value=sanitize_excel(item.codigo_cpc or ""))
            ws4.cell(row=row_num, column=4, value=sanitize_excel(item.descripcion or ""))
            ws4.cell(row=row_num, column=5, value=item.cantidad)
            ws4.cell(row=row_num, column=6, value=sanitize_excel(item.unidad or ""))
            ws4.cell(row=row_num, column=7, value=item.precio_unitario)
            ws4.cell(row=row_num, column=8, value=item.subtotal)
            ws4.cell(row=row_num, column=9, value="Sí" if item.requires_review else "No")
            style_data(ws4, row_num, len(headers4))
            row_num += 1
    auto_width(ws4)

    filepath = os.path.join(EXPORT_DIR, "SupervisorPRO_Export.xlsx")
    wb.save(filepath)
    return filepath
