from typing import List, Optional, Dict
from sqlalchemy.orm import Session
from sqlalchemy import func
from models import Provider, Record
from services.security import sanitize_excel

def get_provider_report(db: Session, search: Optional[str] = None,
                        orden_filtro: Optional[str] = None,
                        fecha_desde: Optional[str] = None,
                        fecha_hasta: Optional[str] = None,
                        page: int = 1, per_page: int = 20) -> Dict:
    subq = db.query(
        Record.provider_id,
        func.group_concat(Record.numero_orden, ", ").label("ordenes"),
        func.group_concat(Record.objeto_contratacion, " | ").label("objetos")
    ).filter(Record.numero_orden.isnot(None), Record.numero_orden != ""
    ).group_by(Record.provider_id).subquery()

    query = db.query(
        Provider.id,
        Provider.nombre,
        Provider.ruc,
        func.count(Record.id).label("total_ordenes"),
        subq.c.ordenes,
        subq.c.objetos,
    ).outerjoin(Record, Record.provider_id == Provider.id
    ).outerjoin(subq, subq.c.provider_id == Provider.id)

    if search:
        term = f"%{search}%"
        query = query.filter(
            Provider.nombre.ilike(term) |
            Provider.ruc.ilike(term) |
            subq.c.ordenes.ilike(term) |
            subq.c.objetos.ilike(term)
        )
    if orden_filtro:
        query = query.filter(subq.c.ordenes.ilike(f"%{orden_filtro}%"))
    if fecha_desde:
        query = query.filter(Record.fecha >= fecha_desde)
    if fecha_hasta:
        query = query.filter(Record.fecha <= fecha_hasta)

    query = query.group_by(Provider.id, Provider.nombre, Provider.ruc, subq.c.ordenes, subq.c.objetos)
    query = query.order_by(Provider.nombre)

    total = query.count()
    rows = query.offset((page - 1) * per_page).limit(per_page).all()

    items = []
    for r in rows:
        items.append({
            "id": r[0],
            "nombre": r[1] or "",
            "ruc": r[2] or "",
            "ordenes": r[4] or "",
            "objeto": r[5] or "",
            "total_infimas": r[3],
        })

    stats = _compute_stats(db)

    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page),
        "stats": stats,
    }

def _compute_stats(db: Session) -> Dict:
    total_prov = db.query(func.count(Provider.id)).scalar() or 0
    total_inf = db.query(func.count(Record.id)).scalar() or 0

    top = db.query(
        Provider.nombre,
        func.count(Record.id).label("cnt")
    ).join(Record, Record.provider_id == Provider.id
    ).group_by(Provider.id, Provider.nombre
    ).order_by(func.count(Record.id).desc()).first()

    return {
        "total_proveedores": total_prov,
        "total_infimas": total_inf,
        "top_proveedor": top[0] if top else "---",
        "top_cantidad": top[1] if top else 0,
    }

def generate_provider_excel(db: Session, search: Optional[str] = None,
                             orden_filtro: Optional[str] = None) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import os, sys, datetime

    subq = db.query(
        Record.provider_id,
        func.group_concat(Record.numero_orden, ", ").label("ordenes"),
        func.group_concat(Record.objeto_contratacion, " | ").label("objetos")
    ).filter(Record.numero_orden.isnot(None), Record.numero_orden != ""
    ).group_by(Record.provider_id).subquery()

    query = db.query(
        Provider.nombre,
        Provider.ruc,
        func.count(Record.id).label("total_ordenes"),
        subq.c.ordenes,
        subq.c.objetos,
    ).outerjoin(Record, Record.provider_id == Provider.id
    ).outerjoin(subq, subq.c.provider_id == Provider.id)

    if search:
        term = f"%{search}%"
        query = query.filter(Provider.nombre.ilike(term) | Provider.ruc.ilike(term) | subq.c.ordenes.ilike(term) | subq.c.objetos.ilike(term))
    if orden_filtro:
        query = query.filter(subq.c.ordenes.ilike(f"%{orden_filtro}%"))

    query = query.group_by(Provider.id, Provider.nombre, Provider.ruc, subq.c.ordenes, subq.c.objetos).order_by(Provider.nombre)
    rows = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Proveedores"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["Nombre Proveedor", "RUC", "N° Orden", "Objeto", "Número de Ínfimas Contratadas"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for ri, row in enumerate(rows, 2):
        vals = [row[0] or "", row[1] or "", row[3] or "", row[4] or "", row[2]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=sanitize_excel(v))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    widths = [35, 20, 30, 50, 25]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.auto_filter.ref = f"A1:E{len(rows) + 1}"

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.cell(row=len(rows) + 3, column=1, value=f"Reporte generado el: {now}").font = Font(italic=True, color="666666")

    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        base = os.path.dirname(base)
    out_dir = os.path.join(base, "exports")
    os.makedirs(out_dir, exist_ok=True)
    filepath = os.path.join(out_dir, "Reporte_Proveedores.xlsx")
    wb.save(filepath)
    return filepath
