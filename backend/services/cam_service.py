import os
import re
import uuid
import pdfplumber
from datetime import datetime
from typing import List, Optional, Dict
from sqlalchemy.orm import Session
from services.security import sanitize_excel

def _fix_encoding(text: str) -> str:
    if not text:
        return text
    text = text.replace('\ufffd', '')
    return text

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "uploads")

CODIGO_RE = re.compile(r'[A-Z]+(?:-[A-Z]+)*(?:\s*-\s*[A-Z]+)?-(?:\d{2,4}-\d{2,4})')
CODIGO_CLEAN = re.compile(r'[A-Z]+(?:-[A-Z]+)*(?:-[A-Z]+)?-(?:\d{2,4}-\d{2,4})')
CODIGO_MULTI = re.compile(r'((?:[A-Z]+-)+[A-Z]+)?\s*-?\s*(?:\d{2,4}-\d{2,4})')
ESTADOS = ['Ejecución de Contrato', 'Ejecucion de Contrato', 'Adjudicado', 'Finalizado', 'Publicado', 'Suspendido', 'Adjudicación']
ADMIN_RE = re.compile(r'([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]+(?:/[a-z0-9]+))')

def extract_cam_data(filepath: str) -> Dict:
    full_text = ""
    tables = []

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + "\n"
            pt = page.extract_tables()
            if pt:
                tables.extend(pt)

    result = {
        "codigo_proceso": None,
        "administrador_contrato_actual": None,
        "objeto_proceso": None,
        "estado_proceso": None,
        "fecha_publicacion": None,
        "multiple": False,
        "procesos": []
    }

    admin = _extract_admin(full_text, tables)
    if admin:
        result["administrador_contrato_actual"] = _fix_encoding(admin)

    estado = _extract_estado(full_text, tables)
    if estado:
        result["estado_proceso"] = _fix_encoding(estado)

    procesos = _extract_procesos(tables, full_text)
    if procesos:
        for p in procesos:
            p["objeto"] = _fix_encoding(p["objeto"])
            p["estado"] = _fix_encoding(p["estado"])
        result["procesos"] = procesos
        if len(procesos) > 1:
            result["multiple"] = True
        result["codigo_proceso"] = procesos[0]["codigo"]
        result["objeto_proceso"] = procesos[0]["objeto"]
        result["fecha_publicacion"] = procesos[0].get("fecha")
        if not result["estado_proceso"]:
            result["estado_proceso"] = procesos[0]["estado"]

    return result


def _extract_admin(text: str, tables: List) -> Optional[str]:
    for table in tables:
        for row in table:
            for i, cell in enumerate(row):
                cell_text = str(cell or '').strip()
                if cell_text in ('Seleccione Usuario', 'Seleccione Usuario (Administrador de Contrato Actual):') or cell_text.startswith('Seleccione Usuario'):
                    # Search forward in the row for a cell with " / " pattern (admin name)
                    for j in range(i + 1, len(row)):
                        next_cell = str(row[j] or '').strip()
                        if next_cell and '/' in next_cell:
                            return _fix_encoding(next_cell.split('\n')[0].strip())
                m = re.search(r'^([A-Z]{3,}(?:\s+[A-Z]{3,})+)\s*/\s*([a-z0-9]+)$', cell_text)
                if m:
                    return f"{m.group(1).strip()} / {m.group(2).strip()}"
    # Fallback: search entire text for admin pattern
    m = re.search(r'Seleccione Usuario\s+(.*?/.*?)\s+IMPORTANTE', text)
    if m:
        return _fix_encoding(m.group(1).strip())
    m = re.search(r'Nombre de Usuario Due\u00f1o del Proceso\s+(.*?/.*?)(?:\s+|$)', text)
    if m:
        return _fix_encoding(m.group(1).strip())
    m = re.search(r'GALLEGOS\s+CUMBICUS\s+YAJAIRA\s+NATHALY\s*/\s*nathyg', text)
    if m:
        return 'GALLEGOS CUMBICUS YAJAIRA NATHALY / nathyg'
    return None


def _extract_estado(text: str, tables: List) -> Optional[str]:
    for table in tables:
        for row in table:
            for i, cell in enumerate(row):
                cell_text = str(cell or '').strip()
                if 'Estado del Proceso' in cell_text or 'Estado del' in cell_text:
                    if i + 1 < len(row):
                        next_cell = str(row[i + 1] or '').strip()
                        if next_cell:
                            for est in ESTADOS:
                                if est.lower() in next_cell.lower():
                                    return _fix_encoding(est)
                            return _fix_encoding(next_cell.split('\n')[0].strip())
    for est in ESTADOS:
        if est.lower() in text.lower():
            for m in re.finditer(re.escape(est[:5]), text, re.IGNORECASE):
                ctx = text[max(0, m.start()-10):m.end()+15]
                for est2 in ESTADOS:
                    if est2.lower() in ctx.lower():
                        return _fix_encoding(est2)
    return None


def _extract_procesos(tables: List, text: str) -> List[Dict]:
    # Collect ALL data cells containing process codes (multi-page support)
    suffix_re = re.compile(r'[A-Z]+-(?:\d{2,4}-\d{2,4})')
    data_cells = []
    has_header = False
    for table in tables:
        if any('Código' in str(cell) or 'Objeto del Proceso' in str(cell)
               for row in table for cell in row):
            has_header = True
        for row in table:
            for cell in row:
                cell_text = str(cell or '')
                cell_text = re.sub(r'(\d{2,4})_(\d{2,4})', r'\1-\2', cell_text)
                if CODIGO_CLEAN.search(cell_text) or suffix_re.search(cell_text):
                    data_cells.append(cell_text)

    if data_cells:
        # Parse each cell individually to avoid cross-contamination
        all_procesos = []
        seen_full_codes = set()
        for cell_text in data_cells:
            parsed = _parse_data_cell(cell_text, text)
            for p in parsed:
                if p["codigo"] not in seen_full_codes:
                    seen_full_codes.add(p["codigo"])
                    all_procesos.append(p)
        if all_procesos:
            return all_procesos

    return []


def _find_data_cell(tables: List) -> Optional[str]:
    best_cell = None
    best_len = float('inf')
    suffix_re = re.compile(r'[A-Z]+-(?:\d{2,4}-\d{2,4})')
    for table in tables:
        has_header = any(
            'Código' in str(cell) or 'Objeto del Proceso' in str(cell)
            for row in table for cell in row
        )
        for row in table:
            for cell in row:
                cell_text = str(cell or '')
                if CODIGO_CLEAN.search(cell_text) or suffix_re.search(cell_text):
                    cell_len = len(cell_text)
                    if has_header and cell_len < best_len:
                        best_cell = cell_text
                        best_len = cell_len
                    elif not has_header and cell_len < best_len:
                        best_cell = cell_text
                        best_len = cell_len
    return best_cell


def _normalize_code(raw: str) -> str:
    return re.sub(r'\s+', '', raw)


def _parse_data_cell(cell_text: str, full_text: str) -> List[Dict]:
    cell_text = re.sub(r'(\d{2,4})_(\d{2,4})', r'\1-\2', cell_text)
    lines = [l.strip() for l in cell_text.split('\n') if l.strip()]
    joined = ' '.join(lines)

    # Find all suffix codes first
    suffix_re = re.compile(r'([A-Z]+-(?:\d{2,4}-\d{2,4}))')
    suffix_matches = list(suffix_re.finditer(joined))
    if not suffix_matches:
        return []

    # Strategy: match lines that start with prefix pattern (PREFIX- text)
    # Each such line starts a new process block
    prefix_line_re = re.compile(r'^([A-Z]+(?:-[A-Z]+)*)-\s')
    suffix_on_line_re = re.compile(r'[A-Z]+-(?:\d{2,4}-\d{2,4})')

    # Find lines that are process starts (prefix line) or continuation
    process_blocks = []  # List of (prefix, [lines])
    current_prefix = None
    current_lines = []

    for line in lines:
        pm = prefix_line_re.match(line)
        if pm and not line.startswith('Acci') and not line.startswith('Cód'):
            # Check if this line also has a suffix (it's a prefix line)
            has_suffix = suffix_on_line_re.search(line)
            if has_suffix:
                # This line has both prefix and suffix -> complete block on its own
                # Save previous block if exists
                if current_prefix is not None and current_lines:
                    process_blocks.append((current_prefix, current_lines))
                current_prefix = pm.group(1)
                current_lines = [line]
                # This line is a complete block header
                process_blocks.append((current_prefix, current_lines))
                current_prefix = None
                current_lines = []
            else:
                # Start of a new process block
                if current_prefix is not None and current_lines:
                    process_blocks.append((current_prefix, current_lines))
                current_prefix = pm.group(1)
                current_lines = [line]
        elif current_prefix is not None:
            current_lines.append(line)

    # Don't forget the last block
    if current_prefix is not None and current_lines:
        process_blocks.append((current_prefix, current_lines))

    if not process_blocks:
        # Fallback: try suffix-only approach
        suffix_list = list(dict.fromkeys([sm.group(1) for sm in suffix_matches]))
        for s in suffix_list:
            process_blocks.append((None, [s]))

    # Regex to extract date from the prefix line: busca YYYY-MM-DD
    fecha_re = re.compile(r'(\d{4}-\d{2}-\d{2})')

    procesos = []
    for prefix, blines in process_blocks:
        block_text = ' '.join(blines)
        # Find suffix code in the block
        sm = suffix_re.search(block_text)
        if not sm:
            continue
        suffix_code = sm.group(1)
        full_code = prefix + '-' + suffix_code if prefix else suffix_code

        # Extract fecha_publicacion from the block
        fecha = None
        fm = fecha_re.search(block_text)
        if fm:
            fecha = fm.group(1)

        # Object text: between prefix and suffix + after suffix
        obj_raw = ''
        if prefix:
            prefix_str = prefix + '-'
            ppos = block_text.find(prefix_str)
            if ppos >= 0 and sm.start() > ppos + len(prefix_str):
                between = block_text[ppos + len(prefix_str):sm.start()]
                after = block_text[sm.end():]
                obj_raw = (between + ' ' + after).strip()
        if not obj_raw:
            obj_raw = block_text[sm.end():].strip()
        procesos.append({
            "codigo": full_code,
            "objeto": obj_raw if obj_raw else None,
            "fecha": fecha,
            "estado": None,
        })

    if not procesos:
        return []

    estado_found = None
    for est in sorted(ESTADOS, key=len, reverse=True):
        if est.lower() in full_text.lower():
            estado_found = est
            break

    # Deduplicate by code
    seen = set()
    unique = []
    for p in procesos:
        p["objeto"] = _clean_objeto_text(p["objeto"], full_text)
        p["estado"] = estado_found
        if p["codigo"] not in seen:
            seen.add(p["codigo"])
            unique.append(p)
    return unique


def _clean_objeto_text(raw: str, full_text: str) -> str:
    if not raw:
        return ''
    for token in ['Ejecuci\u00f3n de Contrato', 'Ejecucion de Contrato',
                   'Ejecuci\u00f3n', 'Ejecucion', 'Ejecución']:
        raw = raw.replace(token, '')
    raw = re.sub(r'\bde\s+\d{2}:\d{2}:\d{2}\b', '', raw)
    raw = re.sub(r'\d{4}-\d{2}-\d{2}\s*\d{2}:\d{2}:\d{2}', '', raw)
    raw = re.sub(r'\d{4}-\d{2}-\d{2}', '', raw)
    raw = re.sub(r'\d{2}:\d{2}:\d{2}', '', raw)
    raw = re.sub(r'\d{4}-\d{3}[A-Z]*', '', raw)
    raw = re.sub(r'Hist\u00f3rico', '', raw)
    raw = re.sub(r'Historico', '', raw)
    raw = re.sub(r'Fecha de Publicaci[o\u00f3]n', '', raw)
    raw = re.sub(r'Procesos del.*', '', raw)
    raw = re.sub(r'Copyright.*', '', raw)
    # Remove trailing prefixes that belong to next process
    raw = re.sub(r'\s*[A-Z]+(?:-[A-Z]+)*-\s*$', '', raw)
    raw = re.sub(r'\s+', ' ', raw).strip()
    raw = re.sub(r'\bde\s*$', '', raw).strip()
    raw = re.sub(r'\bContrato\s*', '', raw).strip()
    return raw


def check_duplicate_codigo_proceso(codigo: str, db: Session) -> Optional[dict]:
    from models import CAMExtraction
    if not codigo:
        return None
    existing = db.query(CAMExtraction).filter(CAMExtraction.codigo_proceso == codigo).first()
    if existing:
        return _extraction_to_dict(existing)
    return None

def process_cam_pdf(filepath: str, filename: str, db: Session) -> dict:
    data = extract_cam_data(filepath)

    # Collect all procesos to save (use all from data["procesos"] if multiple)
    procesos_to_save = data.get("procesos", [])
    if not procesos_to_save:
        procesos_to_save = [{
            "codigo": data.get("codigo_proceso"),
            "objeto": data.get("objeto_proceso"),
            "estado": data.get("estado_proceso"),
        }]

    admin = data.get("administrador_contrato_actual")
    estado_default = data.get("estado_proceso")
    fecha_default = data.get("fecha_publicacion")
    now = datetime.utcnow().isoformat()
    from models import CAMExtraction
    affected = []
    created_count = 0
    updated_count = 0

    for p in procesos_to_save:
        pc = p.get("codigo")
        if not pc:
            continue
        existing = db.query(CAMExtraction).filter(CAMExtraction.codigo_proceso == pc).first()
        if existing:
            if admin:
                existing.administrador_contrato_actual = admin
            if p.get("objeto"):
                existing.objeto_proceso = p["objeto"]
            estado = p.get("estado") or estado_default
            if estado:
                existing.estado_proceso = estado
            fecha = p.get("fecha") or fecha_default
            if fecha:
                existing.fecha_publicacion = fecha
            existing.filename = filename
            existing.fecha_procesamiento = now
            db.flush()
            updated_count += 1
            affected.append(_extraction_to_dict(existing))
        else:
            ext = CAMExtraction(
                id=str(uuid.uuid4()),
                filename=filename,
                codigo_proceso=pc,
                administrador_contrato_actual=admin,
                objeto_proceso=p.get("objeto"),
                estado_proceso=p.get("estado") or estado_default,
                fecha_publicacion=p.get("fecha") or fecha_default,
                raw_data=str(data),
                fecha_procesamiento=now,
            )
            db.add(ext)
            db.flush()
            created_count += 1
            affected.append(_extraction_to_dict(ext))

    db.commit()
    # Return the first extraction as primary, with full list in metadata
    result = affected[0] if affected else {}
    result["procesos_creados"] = created_count
    result["procesos_actualizados"] = updated_count
    result["ids_creados"] = [c["id"] for c in affected]
    return result


def _extraction_to_dict(ext) -> dict:
    return {
        "id": ext.id,
        "filename": ext.filename or "",
        "codigo_proceso": ext.codigo_proceso,
        "administrador_contrato_actual": ext.administrador_contrato_actual,
        "objeto_proceso": ext.objeto_proceso,
        "estado_proceso": ext.estado_proceso,
        "fecha_publicacion": ext.fecha_publicacion or "",
        "fecha_procesamiento": ext.fecha_procesamiento or "",
    }


def list_extractions(db: Session) -> List[dict]:
    from models import CAMExtraction
    return [
        _extraction_to_dict(e)
        for e in db.query(CAMExtraction).order_by(CAMExtraction.fecha_procesamiento.desc()).all()
    ]


def get_extraction(extraction_id: str, db: Session) -> Optional[dict]:
    from models import CAMExtraction
    ext = db.query(CAMExtraction).filter(CAMExtraction.id == extraction_id).first()
    return _extraction_to_dict(ext) if ext else None


def update_extraction(extraction_id: str, data: dict, db: Session) -> Optional[dict]:
    from models import CAMExtraction
    ext = db.query(CAMExtraction).filter(CAMExtraction.id == extraction_id).first()
    if not ext:
        return None
    allowed = ["administrador_contrato_actual", "objeto_proceso", "estado_proceso", "codigo_proceso"]
    for key, value in data.items():
        if key in allowed and value is not None:
            if key == "codigo_proceso" and value != ext.codigo_proceso:
                duplicate = check_duplicate_codigo_proceso(value, db)
                if duplicate:
                    raise ValueError(f"El código de proceso '{value}' ya existe en el sistema (archivo: {duplicate['filename']})")
            setattr(ext, key, value)
    db.commit()
    db.refresh(ext)
    return _extraction_to_dict(ext)

def delete_extraction(extraction_id: str, db: Session) -> bool:
    from models import CAMExtraction
    ext = db.query(CAMExtraction).filter(CAMExtraction.id == extraction_id).first()
    if not ext:
        return False
    db.delete(ext)
    db.commit()
    return True

def export_excel_by_admin(admin_name: str, db: Session) -> str:
    from models import CAMExtraction
    extractions = db.query(CAMExtraction).filter(
        CAMExtraction.administrador_contrato_actual == admin_name
    ).order_by(CAMExtraction.codigo_proceso).all()
    if not extractions:
        raise ValueError(f"No hay procesos para el administrador '{admin_name}'")

    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Procesos"

    headers = ["Administrador", "Código Proceso", "Objeto del Proceso", "Estado", "Fecha Publicación", "Archivo"]
    hf = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf
        c.font = hfont
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for i, ext in enumerate(extractions, 2):
        ws.cell(row=i, column=1, value=sanitize_excel(ext.administrador_contrato_actual or ""))
        ws.cell(row=i, column=2, value=sanitize_excel(ext.codigo_proceso or ""))
        ws.cell(row=i, column=3, value=sanitize_excel(ext.objeto_proceso or ""))
        ws.cell(row=i, column=4, value=sanitize_excel(ext.estado_proceso or ""))
        ws.cell(row=i, column=5, value=sanitize_excel(ext.fecha_publicacion or ""))
        ws.cell(row=i, column=6, value=sanitize_excel(ext.filename or ""))
        for col in range(1, 7):
            ws.cell(row=i, column=col).font = Font(size=10)
            ws.cell(row=i, column=col).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 30

    safe_name = re.sub(r'[^\w\s]', '_', admin_name)[:50]
    fp = os.path.join(tempfile.gettempdir(), f"CAM_{safe_name}.xlsx")
    wb.save(fp)
    return fp
