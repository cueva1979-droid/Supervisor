import os
import sys
import io
import uuid
import json
import re
import secrets
from datetime import datetime as dt
from collections import defaultdict
from contextlib import asynccontextmanager
from typing import List, Optional

import uvicorn
from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Query, Request, Response as FastResponse
from fastapi.responses import FileResponse
from fastapi.security import HTTPBearer
from sqlalchemy.orm import Session
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from database import init_db, get_db, SessionLocal, create_default_admin
from models import (
    Record, Item, Provider, Base, PACDocument, PACCertificate,
    CPCCatalog, CPCLoadedData, CEExtractionDB, CEItemDB, CAMExtraction,
    User, AuditLog, LoginAttempt
)
from schemas import (
    RecordResponse, RecordCreate, RecordUpdate,
    ProviderResponse, ProviderCreate, ProviderUpdate,
    ItemResponse, DashboardResponse,
    PACDocumentResponse, PACDocumentUpdate,
    PACCertificateResponse, PACCertificateCreate,
    CPCCatalogResponse, CPCCatalogCreate,
    CPCLoadedDataResponse, CPCLoadedDataCreate,
    PACAnalysisResponse,
    LoginRequest, LoginResponse, RefreshRequest,
    UserCreate, UserUpdate, UserResponse,
    AuditLogResponse, ChangePasswordRequest,
)
from services.extraction_service import upload_and_process, delete_record, get_records
from services.provider_service import (
    get_providers, get_provider, create_provider,
    update_provider, delete_provider, count_provider_records
)
from services.excel_service import generate_excel
from services.report_service import get_provider_report, generate_provider_excel
from services.product_report_service import get_product_report, generate_product_excel
from services.pac_service import extract_excel_data, extract_pdf_data, extract_pdf_tabular_data, classify_period
from services.pac_certificate_service import generate_certificate_docx, generate_certificate_from_template
from services.ce_service import process_pdf, list_extractions, get_extraction, delete_extraction, clear_all, export_excel, export_excel_by_admin
from services.cam_service import process_cam_pdf, list_extractions as cam_list_extractions, get_extraction as cam_get_extraction, delete_extraction as cam_delete_extraction
from extractor_oc.parser import PDFExtractor
from config import settings
from auth import (
    hash_password, verify_password, create_access_token, create_refresh_token,
    decode_token, get_current_user, require_auth, require_role, require_module,
    check_login_rate_limit, record_login_attempt, log_audit, sanitize_input,
    generate_csrf_token, validate_csrf,
)

security = HTTPBearer(auto_error=False)

# --- CORS + Security Headers Middleware (pure ASGI) ---
CORS_ALLOW_ORIGINS = settings.ALLOWED_ORIGINS
CORS_ALLOW_METHODS = b"GET, POST, PUT, DELETE, OPTIONS, PATCH"
CORS_ALLOW_HEADERS = b"Authorization, Content-Type, X-CSRF-Token, X-Requested-With, Accept, Origin"
CORS_EXPOSE_HEADERS = b"Content-Disposition, Content-Type"

SECURITY_HEADERS = [
    (b"x-content-type-options", b"nosniff"),
    (b"x-frame-options", b"DENY"),
    (b"x-xss-protection", b"1; mode=block"),
    (b"strict-transport-security", b"max-age=31536000; includeSubDomains"),
    (b"cache-control", b"no-store, no-cache, must-revalidate"),
    (b"pragma", b"no-cache"),
    (b"content-security-policy", b"default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'"),
    (b"referrer-policy", b"strict-origin-when-cross-origin"),
    (b"permissions-policy", b"camera=(), microphone=(), geolocation=()"),
]

class CORSAndSecurityMiddleware:
    def __init__(self, app):
        self.app = app

    async def __call__(self, scope, receive, send):
        if scope["type"] != "http":
            await self.app(scope, receive, send)
            return

        # Strip the /api prefix so the frontend can call same-origin routes
        path = scope.get("path", "")
        if path.startswith("/api") and (len(path) == 4 or path[4] == "/"):
            scope["path"] = path[4:] or "/"

        method = scope.get("method", "")
        headers = dict(scope.get("headers", []))
        origin = headers.get(b"origin", b"")

        # Handle OPTIONS preflight
        if method == "OPTIONS":
            allowed = origin and origin.decode() in CORS_ALLOW_ORIGINS
            resp_headers = []
            if allowed:
                resp_headers = [
                    (b"access-control-allow-origin", origin),
                    (b"access-control-allow-methods", CORS_ALLOW_METHODS),
                    (b"access-control-allow-headers", CORS_ALLOW_HEADERS),
                    (b"access-control-allow-credentials", b"true"),
                    (b"access-control-max-age", b"3600"),
                    (b"vary", b"Origin"),
                ]
            resp_headers.append((b"content-length", b"0"))
            await send({"type": "http.response.start", "status": 200, "headers": resp_headers})
            await send({"type": "http.response.body", "body": b""})
            return

        async def send_with_headers(message):
            if message["type"] == "http.response.start":
                h = list(message.get("headers", []))
                # Add CORS header for browser requests
                if origin:
                    h.append((b"access-control-allow-origin", origin))
                    h.append((b"access-control-allow-credentials", b"true"))
                    h.append((b"access-control-expose-headers", CORS_EXPOSE_HEADERS))
                    h.append((b"vary", b"Origin"))
                # Add security headers
                h.extend(SECURITY_HEADERS)
                message["headers"] = h
            await send(message)

        await self.app(scope, receive, send_with_headers)

@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    db = SessionLocal()
    try:
        create_default_admin(db)
    finally:
        db.close()
    from services.backup_service import start_scheduler
    start_scheduler()
    yield

app = FastAPI(title=settings.APP_NAME, version=settings.VERSION, lifespan=lifespan)

app.add_middleware(CORSAndSecurityMiddleware)

@app.get("/health")
def health():
    return {"status": "ok", "app": settings.APP_NAME, "version": settings.VERSION}

# ==================== Auth Endpoints ====================

@app.post("/auth/login", response_model=LoginResponse)
def login(data: LoginRequest, request: Request, db: Session = Depends(get_db)):
    ip = request.client.host if request.client else "unknown"
    check_login_rate_limit(data.username, ip, db)
    user = db.query(User).filter(User.username == data.username).first()
    if not user or not verify_password(data.password, user.password_hash):
        record_login_attempt(data.username, ip, False, db)
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Cuenta desactivada")
    user.last_login = datetime.utcnow()
    db.commit()
    record_login_attempt(data.username, ip, True, db)
    access_token = create_access_token(user.id, user.role)
    refresh_token = create_refresh_token(user.id)
    log_audit(user.id, "LOGIN", "auth", details="Inicio de sesión exitoso", ip_address=ip, db=db)
    return LoginResponse(
        access_token=access_token,
        refresh_token=refresh_token,
        user={"id": user.id, "username": user.username, "role": user.role, "email": user.email or ""},
    )

@app.post("/auth/refresh")
def refresh_token(data: RefreshRequest, db: Session = Depends(get_db)):
    payload = decode_token(data.refresh_token)
    if not payload or payload.get("type") != "refresh":
        raise HTTPException(status_code=401, detail="Token de actualización inválido")
    user = db.query(User).filter(User.id == int(payload["sub"])).first()
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="Usuario no válido")
    access_token = create_access_token(user.id, user.role)
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/auth/me", response_model=UserResponse)
def get_me(user: User = Depends(require_auth)):
    return user

@app.put("/auth/password")
def change_password(data: ChangePasswordRequest, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    if not verify_password(data.current_password, user.password_hash):
        raise HTTPException(status_code=400, detail="Contraseña actual incorrecta")
    if len(data.new_password) < 8:
        raise HTTPException(status_code=400, detail="La nueva contraseña debe tener al menos 8 caracteres")
    user.password_hash = hash_password(data.new_password)
    db.commit()
    log_audit(user.id, "PASSWORD_CHANGE", "auth", details="Cambio de contraseña", db=db)
    return {"message": "Contraseña actualizada exitosamente"}

# ==================== User Management (Admin only) ====================

@app.get("/admin/users", response_model=List[UserResponse])
def list_users(user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    return db.query(User).order_by(User.username).all()

@app.post("/admin/users", response_model=UserResponse)
def create_user(data: UserCreate, user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    if len(data.password) < 8:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 8 caracteres")
    existing = db.query(User).filter(User.username == data.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="El usuario ya existe")
    new_user = User(
        username=data.username,
        email=data.email,
        password_hash=hash_password(data.password),
        role=data.role if data.role in settings.ROLES else "operator",
        is_active=True,
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    log_audit(user.id, "USER_CREATE", "users", resource_id=str(new_user.id), details=f"Creó usuario: {data.username}", db=db)
    return new_user

@app.put("/admin/users/{user_id}", response_model=UserResponse)
def update_user(user_id: int, data: UserUpdate, user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if data.username is not None:
        target.username = data.username
    if data.email is not None:
        target.email = data.email
    if data.password is not None:
        if len(data.password) < 8:
            raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 8 caracteres")
        target.password_hash = hash_password(data.password)
    if data.role is not None:
        if data.role not in settings.ROLES:
            raise HTTPException(status_code=400, detail=f"Rol inválido. Roles válidos: {list(settings.ROLES.keys())}")
        target.role = data.role
    if data.is_active is not None:
        target.is_active = data.is_active
    db.commit()
    db.refresh(target)
    log_audit(user.id, "USER_UPDATE", "users", resource_id=str(user_id), details=f"Actualizó usuario {target.username}", db=db)
    return target

@app.delete("/admin/users/{user_id}")
def delete_user(user_id: int, user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    if user_id == user.id:
        raise HTTPException(status_code=400, detail="No puede eliminarse a sí mismo")
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    db.delete(target)
    db.commit()
    log_audit(user.id, "USER_DELETE", "users", resource_id=str(user_id), details=f"Eliminó usuario {target.username}", db=db)
    return {"message": "Usuario eliminado"}

# ==================== Audit Logs (Admin only) ====================

@app.get("/admin/audit-logs", response_model=List[AuditLogResponse])
def list_audit_logs(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    user: User = Depends(require_role("admin")),
    db: Session = Depends(get_db),
):
    query = db.query(AuditLog).order_by(AuditLog.timestamp.desc())
    total = query.count()
    logs = query.offset((page - 1) * per_page).limit(per_page).all()
    return logs

# ==================== Protected Endpoints ====================

@app.post("/upload", response_model=List[dict])
def upload_files(files: List[UploadFile] = File(...), user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    if not files:
        raise HTTPException(status_code=400, detail="No se enviaron archivos")
    results = upload_and_process(files, db)
    response = []
    for filename, record, error in results:
        if error:
            response.append({"filename": filename, "status": "error", "error": error})
        else:
            response.append({
                "filename": filename,
                "status": "success",
                "record": RecordResponse.model_validate(record).model_dump()
            })
    return response

@app.get("/records", response_model=List[RecordResponse])
def list_records(search: Optional[str] = Query(None), user: User = Depends(require_auth), db: Session = Depends(get_db)):
    records = get_records(db, search)
    return [RecordResponse.model_validate(r).model_dump() for r in records]

@app.get("/records/{record_id}", response_model=RecordResponse)
def get_record(record_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    record = db.query(Record).filter(Record.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    return RecordResponse.model_validate(record).model_dump()

@app.put("/records/{record_id}", response_model=RecordResponse)
def update_record(record_id: int, data: RecordUpdate, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    record = db.query(Record).filter(Record.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    update_data = data.model_dump(exclude={"items"}, exclude_unset=True)

    if "numero_orden" in update_data and update_data["numero_orden"]:
        existing_ce = db.query(CEExtractionDB).filter(
            CEExtractionDB.orden_compra == update_data["numero_orden"]
        ).first()
        if existing_ce:
            raise HTTPException(
                status_code=400,
                detail=f"El número de orden '{update_data['numero_orden']}' ya existe en el Catálogo Electrónico "
                       f"(archivo: {existing_ce.filename})"
            )

    for key, value in update_data.items():
        setattr(record, key, value)
    if data.items is not None:
        db.query(Item).filter(Item.record_id == record_id).delete()
        for item_data in data.items:
            item = Item(
                record_id=record_id,
                codigo_cpc=item_data.codigo_cpc,
                descripcion=item_data.descripcion,
                cantidad=item_data.cantidad,
                unidad=item_data.unidad,
                precio_unitario=item_data.precio_unitario,
                subtotal=item_data.cantidad * item_data.precio_unitario,
                requires_review=item_data.requires_review,
            )
            db.add(item)
        total = sum(item.cantidad * item.precio_unitario for item in data.items)
        record.monto_total = round(total, 2)
    db.commit()
    db.refresh(record)
    return RecordResponse.model_validate(record).model_dump()

@app.delete("/records/{record_id}")
def remove_record(record_id: int, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    success = delete_record(record_id, db)
    if not success:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    log_audit(user.id, "DELETE", "records", resource_id=str(record_id), db=db)
    return {"status": "ok", "message": "Registro eliminado"}

@app.get("/providers", response_model=List[ProviderResponse])
def list_providers(search: Optional[str] = Query(None), user: User = Depends(require_auth), db: Session = Depends(get_db)):
    providers = get_providers(db, search)
    counts = count_provider_records(db)
    result = []
    for p in providers:
        pd = ProviderResponse.model_validate(p).model_dump()
        pd["contratos"] = counts.get(p.id, 0)
        result.append(pd)
    return result

@app.get("/providers/{provider_id}", response_model=ProviderResponse)
def get_provider_endpoint(provider_id: int, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    provider = get_provider(db, provider_id)
    if not provider:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    pd = ProviderResponse.model_validate(provider).model_dump()
    pd["contratos"] = count_provider_records(db).get(provider.id, 0)
    return pd

@app.post("/providers", response_model=ProviderResponse, status_code=201)
def create_provider_endpoint(data: ProviderCreate, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    try:
        provider = create_provider(db, data)
        log_audit(user.id, "CREATE", "providers", resource_id=str(provider.id), db=db)
        return ProviderResponse.model_validate(provider).model_dump()
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.put("/providers/{provider_id}", response_model=ProviderResponse)
def update_provider_endpoint(provider_id: int, data: ProviderUpdate, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    provider = update_provider(db, provider_id, data)
    if not provider:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    log_audit(user.id, "UPDATE", "providers", resource_id=str(provider_id), db=db)
    return ProviderResponse.model_validate(provider).model_dump()

@app.delete("/providers/{provider_id}")
def remove_provider(provider_id: int, user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    success = delete_provider(db, provider_id)
    if not success:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    log_audit(user.id, "DELETE", "providers", resource_id=str(provider_id), db=db)
    return {"status": "ok", "message": "Proveedor eliminado"}

@app.get("/export/excel")
def export_all_excel(user: User = Depends(require_auth), db: Session = Depends(get_db)):
    try:
        filepath = generate_excel(db)
        return FileResponse(
            filepath,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="SupervisorPDF_Export.xlsx"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

@app.get("/export/administradores")
def export_administradores(user: User = Depends(require_auth), db: Session = Depends(get_db)):
    try:
        records = db.query(Record).filter(Record.administrador.isnot(None)).filter(Record.administrador != "").order_by(Record.administrador).all()
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "Administradores"
        headers = ["Administrador", "N° Orden", "Proveedor", "Objeto de Contratación", "Fecha", "Monto"]
        hf = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal="center")
            c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for i, r in enumerate(records, 2):
            ws.cell(row=i, column=1, value=r.administrador)
            ws.cell(row=i, column=2, value=r.numero_orden or "")
            ws.cell(row=i, column=3, value=r.proveedor or "")
            ws.cell(row=i, column=4, value=r.objeto_contratacion or "")
            ws.cell(row=i, column=5, value=r.fecha or "")
            ws.cell(row=i, column=6, value=r.monto_total or 0)
            for col in range(1, 7):
                ws.cell(row=i, column=col).font = Font(size=10)
                ws.cell(row=i, column=col).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        import tempfile
        fp = os.path.join(tempfile.gettempdir(), "Administradores.xlsx")
        wb.save(fp)
        return FileResponse(fp, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="Administradores.xlsx")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

@app.get("/productos")
def list_productos(
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    query = db.query(Item, Record).join(Record, Item.record_id == Record.id)
    if search:
        term = f"%{search}%"
        query = query.filter(
            Item.codigo_cpc.ilike(term) |
            Item.descripcion.ilike(term) |
            Item.unidad.ilike(term) |
            Record.numero_orden.ilike(term) |
            Record.proveedor.ilike(term)
        )
    total = query.count()
    results = query.order_by(Record.fecha_procesamiento.desc(), Item.id).offset((page - 1) * per_page).limit(per_page).all()
    items = []
    for item, record in results:
        items.append({
            "id": item.id,
            "codigo_cpc": item.codigo_cpc or "",
            "descripcion": item.descripcion or "",
            "unidad": item.unidad or "",
            "cantidad": item.cantidad or 0,
            "v_unitario": item.precio_unitario or 0,
            "v_total": item.subtotal or 0,
            "numero_orden": record.numero_orden or "",
            "proveedor": record.proveedor or "",
        })
    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page),
    }

@app.get("/dashboard", response_model=DashboardResponse)
def get_dashboard(user: User = Depends(require_auth), db: Session = Depends(get_db)):
    records = db.query(Record).all()
    providers = db.query(Provider).all()

    total_docs = len(records)
    total_prov = len(providers)
    total_ords = len([r for r in records if r.numero_orden])
    total_montos = sum(r.monto_total or 0 for r in records)
    ultimos = db.query(Record).order_by(Record.fecha_procesamiento.desc()).limit(5).all()

    ordenes_por_mes = defaultdict(int)
    for r in records:
        try:
            fecha_str = r.fecha or ""
            if "/" in fecha_str:
                parts = fecha_str.split("/")
                mes = parts[1] if len(parts) > 1 else "00"
                anio = parts[2] if len(parts) > 2 else "0000"
                key = f"{mes}/{anio}"
            else:
                key = "Sin fecha"
            ordenes_por_mes[key] += 1
        except:
            ordenes_por_mes["Sin fecha"] += 1

    montos_por_proveedor = defaultdict(float)
    for r in records:
        prov = r.proveedor or "Desconocido"
        montos_por_proveedor[prov] += r.monto_total or 0

    return DashboardResponse(
        total_documentos=total_docs,
        total_proveedores=total_prov,
        total_ordenes=total_ords,
        total_montos=round(total_montos, 2),
        ultimos_registros=[RecordResponse.model_validate(r).model_dump() for r in ultimos],
        ordenes_por_mes=dict(ordenes_por_mes),
        montos_por_proveedor={k: round(v, 2) for k, v in montos_por_proveedor.items()}
    )

@app.get("/administradores")
def list_administradores(search: Optional[str] = Query(None), user: User = Depends(require_auth), db: Session = Depends(get_db)):
    query = db.query(Record).filter(Record.administrador.isnot(None)).filter(Record.administrador != "")
    if search:
        query = query.filter(Record.administrador.ilike(f"%{search}%"))
    records = query.order_by(Record.administrador).all()
    result = []
    seen = set()
    for r in records:
        admin = r.administrador.strip() if r.administrador else ""
        key = admin.lower()
        if key not in seen:
            seen.add(key)
            result.append({
                "administrador": admin,
                "numero_orden": r.numero_orden,
                "proveedor": r.proveedor,
                "objeto_contratacion": r.objeto_contratacion,
                "fecha": r.fecha,
                "monto_total": r.monto_total,
                "record_id": r.id,
                "filename": r.filename,
            })
        else:
            for item in result:
                if item["administrador"].lower() == key:
                    if r.numero_orden and r.numero_orden not in item.get("ordenes", []):
                        if "ordenes" not in item:
                            item["ordenes"] = []
                        item["ordenes"].append(r.numero_orden)
                    break
    return result

@app.get("/reports/providers")
def report_providers(
    search: Optional[str] = Query(None),
    orden_filtro: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    return get_provider_report(db, search, orden_filtro, fecha_desde, fecha_hasta, page, per_page)

@app.get("/reports/providers/excel")
def report_providers_excel(
    search: Optional[str] = Query(None),
    orden_filtro: Optional[str] = Query(None),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    try:
        filepath = generate_provider_excel(db, search, orden_filtro)
        return FileResponse(
            filepath,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="Reporte_Proveedores.xlsx"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

@app.get("/reports/ordenes")
def report_ordenes(
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    query = db.query(Record)
    if search:
        term = f"%{search}%"
        query = query.filter(
            Record.numero_orden.ilike(term) |
            Record.objeto_contratacion.ilike(term) |
            Record.proveedor.ilike(term) |
            Record.administrador.ilike(term)
        )
    total = query.count()
    results = query.order_by(Record.fecha_procesamiento.desc()).offset((page - 1) * per_page).limit(per_page).all()
    items = []
    for r in results:
        items.append({
            "numero_orden": r.numero_orden or "",
            "objeto_contratacion": r.objeto_contratacion or "",
            "fecha": r.fecha or "",
            "plazo_entrega": r.plazo_entrega or "",
            "proveedor": r.proveedor or "",
            "administrador": r.administrador or "",
        })
    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page),
    }

@app.get("/reports/products")
def report_products(
    search: Optional[str] = Query(None),
    producto: Optional[str] = Query(None),
    proveedor: Optional[str] = Query(None),
    codigo_cpc: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    precio_min: Optional[float] = Query(None),
    precio_max: Optional[float] = Query(None),
    codigo_proceso: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    return get_product_report(db, search, producto, proveedor, codigo_cpc,
                               fecha_desde, fecha_hasta, precio_min, precio_max,
                               codigo_proceso, page, per_page)

@app.get("/reports/products/excel")
def report_products_excel(
    search: Optional[str] = Query(None),
    producto: Optional[str] = Query(None),
    proveedor: Optional[str] = Query(None),
    codigo_cpc: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    precio_min: Optional[float] = Query(None),
    precio_max: Optional[float] = Query(None),
    codigo_proceso: Optional[str] = Query(None),
    user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    try:
        filepath = generate_product_excel(db, search, producto, proveedor, codigo_cpc,
                                           fecha_desde, fecha_hasta, precio_min, precio_max,
                                           codigo_proceso)
        return FileResponse(
            filepath,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="Reporte_Productos_Precios.xlsx"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")

@app.delete("/reset-db")
def reset_database(user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    try:
        db.query(Item).delete()
        db.query(Record).delete()
        db.query(Provider).delete()
        db.commit()
        return {"status": "ok", "message": "Base de datos reiniciada exitosamente"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al reiniciar base de datos: {str(e)}")


@app.post("/backup")
def create_backup_endpoint(user: User = Depends(require_role("admin"))):
    from services.backup_service import create_backup
    try:
        result = create_backup()
        return {"status": "ok", "backup": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al crear backup: {str(e)}")


@app.get("/backup/info")
def get_backup_info_endpoint(user: User = Depends(require_auth)):
    from services.backup_service import get_backup_info
    return get_backup_info()


@app.post("/backup/auto-toggle")
def toggle_auto_backup(data: dict, user: User = Depends(require_role("admin"))):
    from services.backup_service import set_auto_backup
    enabled = data.get("enabled", False)
    return set_auto_backup(enabled)


@app.post("/backup/restore")
def restore_backup_endpoint(data: dict, user: User = Depends(require_role("admin"))):
    from services.backup_service import restore_backup
    filename = data.get("filename", "")
    if not filename:
        raise HTTPException(status_code=400, detail="Debe especificar un archivo de backup")
    try:
        result = restore_backup(filename)
        return result
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al restaurar backup: {str(e)}")


# ==================== PAC Module Endpoints ====================

@app.get("/pac/documents")
def pac_list_documents(user: User = Depends(require_module("pac")), db: Session = Depends(get_db)):
    docs = db.query(PACDocument).order_by(PACDocument.upload_date.desc()).all()
    return [PACDocumentResponse.model_validate(d).model_dump() for d in docs]


@app.get("/pac/documents/{doc_id}")
def pac_get_document(doc_id: str, user: User = Depends(require_module("pac")), db: Session = Depends(get_db)):
    doc = db.query(PACDocument).filter(PACDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    return PACDocumentResponse.model_validate(d).model_dump()


@app.post("/pac/documents")
async def pac_upload_document(file: UploadFile = File(...), user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    if not file:
        raise HTTPException(status_code=400, detail="No se envió archivo")

    filename = file.filename or "unknown"
    ext = filename.lower().split('.')[-1] if '.' in filename else ""

    contents = await file.read()

    try:
        if ext in ('xlsx', 'xls'):
            documents = extract_excel_data(contents)
        elif ext == 'pdf':
            documents = extract_pdf_tabular_data(contents)
            if not documents:
                doc = extract_pdf_data(contents)
                documents = [doc] if doc else []
        else:
            raise HTTPException(status_code=400, detail="Formato no soportado. Use .xlsx, .xls o .pdf")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer el archivo: {str(e)}")

    if not documents:
        raise HTTPException(status_code=400, detail="No se encontraron registros válidos")

    inserted_ids = []
    updated_count = 0
    inserted_count = 0
    pdf_b64 = None
    if ext == 'pdf':
        import base64
        pdf_b64 = base64.b64encode(contents).decode('utf-8')

    for doc_data in documents:
        partida = doc_data.get("partida_presupuestaria", "")
        cpc = doc_data.get("cpc", "")
        periodo = doc_data.get("periodo", "")

        existing = db.query(PACDocument).filter(
            PACDocument.partida_presupuestaria == partida,
            PACDocument.cpc == cpc,
            PACDocument.periodo == periodo
        ).first()

        if existing:
            existing.filename = filename
            existing.tipo_compra = doc_data.get("tipo_compra")
            existing.tipo_regimen = doc_data.get("tipo_regimen")
            existing.procedimiento = doc_data.get("procedimiento")
            existing.descripcion = doc_data.get("descripcion")
            existing.costo_unitario = doc_data.get("costo_unitario")
            if pdf_b64:
                existing.pdf_data = pdf_b64
            existing.upload_date = dt.utcnow().isoformat()
            updated_count += 1
        else:
            new_doc = PACDocument(
                id=str(uuid.uuid4()),
                filename=filename,
                partida_presupuestaria=partida,
                cpc=cpc,
                tipo_compra=doc_data.get("tipo_compra"),
                tipo_regimen=doc_data.get("tipo_regimen"),
                procedimiento=doc_data.get("procedimiento"),
                descripcion=doc_data.get("descripcion"),
                costo_unitario=doc_data.get("costo_unitario"),
                periodo=periodo,
                pdf_data=pdf_b64,
            )
            db.add(new_doc)
            inserted_count += 1
            inserted_ids.append(new_doc.id)

    db.commit()

    return {
        "message": f"Proceso completado: {inserted_count} nuevos, {updated_count} actualizados",
        "inserted": inserted_count,
        "updated": updated_count,
        "extractedData": documents[:100],
    }


@app.put("/pac/documents/{doc_id}")
def pac_update_document(doc_id: str, data: PACDocumentUpdate, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    doc = db.query(PACDocument).filter(PACDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(doc, key, value)
    db.commit()
    log_audit(user.id, "UPDATE", "pac_documents", resource_id=doc_id, db=db)
    return {"message": "Documento actualizado exitosamente"}


@app.delete("/pac/documents/{doc_id}")
def pac_delete_document(doc_id: str, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    doc = db.query(PACDocument).filter(PACDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    db.delete(doc)
    db.commit()
    log_audit(user.id, "DELETE", "pac_documents", resource_id=doc_id, db=db)
    return {"message": "Documento eliminado"}


@app.delete("/pac/documents")
def pac_delete_all_documents(user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    db.query(PACDocument).delete()
    db.commit()
    log_audit(user.id, "DELETE_ALL", "pac_documents", db=db)
    return {"message": "Todos los documentos han sido eliminados"}


@app.post("/pac/documents/clean-duplicates")
def pac_clean_duplicates(user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    docs = db.query(PACDocument).order_by(PACDocument.partida_presupuestaria, PACDocument.cpc, PACDocument.periodo, PACDocument.upload_date).all()
    seen = set()
    to_delete = []
    for doc in docs:
        key = f"{doc.partida_presupuestaria}|{doc.cpc}|{doc.periodo}".lower()
        if key in seen:
            to_delete.append(doc.id)
        else:
            seen.add(key)
    for doc_id in to_delete:
        d = db.query(PACDocument).filter(PACDocument.id == doc_id).first()
        if d:
            db.delete(d)
    db.commit()
    log_audit(user.id, "CLEAN_DUPS", "pac_documents", details=f"Eliminados {len(to_delete)} duplicados", db=db)
    return {"message": f"Se eliminaron {len(to_delete)} duplicados", "deletedCount": len(to_delete)}


@app.get("/pac/analysis/periods")
def pac_period_analysis(user: User = Depends(require_module("pac")), db: Session = Depends(get_db)): 
    docs = db.query(PACDocument).order_by(PACDocument.periodo).all()
    current_month = dt.now().month
    result = []
    for doc in docs:
        analysis = classify_period(doc.periodo, current_month)
        result.append({
            "id": doc.id,
            "filename": doc.filename,
            "partida_presupuestaria": doc.partida_presupuestaria,
            "cpc": doc.cpc,
            "descripcion": doc.descripcion,
            "costo_unitario": doc.costo_unitario,
            "periodo": doc.periodo,
            "periodCategory": analysis["periodCategory"],
            "status": analysis["status"],
        })
    return result


# Certificate endpoints

@app.get("/pac/certificates")
def pac_list_certificates(user: User = Depends(require_module("pac")), db: Session = Depends(get_db)):
    certs = db.query(PACCertificate).order_by(PACCertificate.created_date.desc()).all()
    return [PACCertificateResponse.model_validate(c).model_dump() for c in certs]


@app.post("/pac/certificates")
def pac_create_certificate(data: PACCertificateCreate, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    cert_id = str(uuid.uuid4())
    cert = PACCertificate(
        id=cert_id,
        document_id=data.document_id,
        certificate_type=data.certificate_type,
        cert_nro=data.cert_nro,
        data=json.dumps(data.data) if data.data else None,
    )
    db.add(cert)
    db.commit()
    log_audit(user.id, "CREATE", "pac_certificates", resource_id=cert_id, db=db)
    return {"id": cert_id, "message": "Certificado creado exitosamente"}


@app.delete("/pac/certificates/all")
def pac_delete_all_certificates(user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    db.query(PACCertificate).delete()
    db.commit()
    log_audit(user.id, "DELETE_ALL", "pac_certificates", db=db)
    return {"message": "Historial de certificados eliminado"}


@app.get("/pac/certificates/generate/{doc_id}")
def pac_generate_certificate(doc_id: str, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    doc = db.query(PACDocument).filter(PACDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")

    # Check for template file
    template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "plantilla.docx")
    if os.path.exists(template_path):
        docx_bytes = generate_certificate_from_template(
            {"partida_presupuestaria": doc.partida_presupuestaria,
             "cpc": doc.cpc, "tipo_compra": doc.tipo_compra,
             "tipo_regimen": doc.tipo_regimen, "procedimiento": doc.procedimiento,
             "descripcion": doc.descripcion, "costo_unitario": doc.costo_unitario,
             "periodo": doc.periodo},
            template_path
        )
    else:
        # Generate programmatic certificate
        docx_bytes = generate_certificate_docx({
            "cert_nro": "S/N",
            "objeto": doc.descripcion or "",
            "partida": doc.partida_presupuestaria or "",
            "cpc": doc.cpc or "",
            "tipo_compra": doc.tipo_compra or "",
            "tipo_regimen": doc.tipo_regimen or "",
            "procedimiento": doc.procedimiento or "",
            "periodo": doc.periodo or "",
            "valor": doc.costo_unitario or 0,
            "verificacion_catalogo": "SI",
            "elaborado_por": "",
            "cargo": "",
            "aprobado_por": "",
            "cargo_aprobado": "",
        })

    if not docx_bytes:
        raise HTTPException(status_code=500, detail="Error al generar el certificado")

    return FastResponse(
        content=docx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename=Certificado_PAC_{doc.cpc}.docx"}
    )


@app.post("/pac/certificates/generate-custom")
def pac_generate_custom_certificate(data: dict, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    try:
        # Check for unique cert_nro
        cert_nro = data.get("cert_nro")
        if cert_nro:
            existing = db.query(PACCertificate).filter(PACCertificate.cert_nro == str(cert_nro)).first()
            if existing:
                raise HTTPException(status_code=400, detail=f"El número de certificación {cert_nro} ya existe")

        docx_bytes = generate_certificate_docx(data)

        if not docx_bytes:
            raise HTTPException(status_code=500, detail="Error al generar el certificado")

        # Save certificate record
        cert_id = str(uuid.uuid4())
        cert = PACCertificate(
            id=cert_id,
            document_id=None,
            certificate_type="MANUAL",
            cert_nro=str(cert_nro) if cert_nro else None,
            data=json.dumps(data),
        )
        db.add(cert)
        db.commit()

        return FastResponse(
            content=docx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f"attachment; filename=Certificacion_PAC_{cert_nro}.docx"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar la certificación: {str(e)}")


@app.get("/pac/template/download")
def pac_download_template(user: User = Depends(require_role("admin", "operator"))):
    template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "plantilla.docx")
    if os.path.exists(template_path):
        return FileResponse(template_path, filename="plantilla_maestra.docx")
    raise HTTPException(status_code=404, detail="Plantilla no encontrada")


# CPC endpoints

@app.get("/pac/cpc/catalog")
def pac_cpc_catalog(user: User = Depends(require_module("pac")), db: Session = Depends(get_db)):
    catalog = db.query(CPCCatalog).order_by(CPCCatalog.cpc).all()
    return [CPCCatalogResponse.model_validate(c).model_dump() for c in catalog]


@app.post("/pac/cpc/bulk")
def pac_cpc_bulk(data: dict, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    items = data.get("items", [])
    if not items or not isinstance(items, list):
        raise HTTPException(status_code=400, detail="Se requiere una lista de items")

    for item in items:
        cpc_code = str(item.get("cpc", "")).strip()
        if not cpc_code:
            continue
        existing = db.query(CPCCatalog).filter(CPCCatalog.cpc == cpc_code).first()
        if existing:
            existing.descripcion = item.get("descripcion", existing.descripcion)
            existing.umbral = item.get("umbral", existing.umbral)
        else:
            new_item = CPCCatalog(
                cpc=cpc_code,
                descripcion=item.get("descripcion", ""),
                umbral=item.get("umbral", 0),
            )
            db.add(new_item)
    db.commit()
    return {"message": f"{len(items)} registros CPC guardados/actualizados exitosamente"}


@app.delete("/pac/cpc/catalog/{cpc_code}")
def pac_cpc_delete(cpc_code: str, user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    item = db.query(CPCCatalog).filter(CPCCatalog.cpc == cpc_code).first()
    if item:
        db.delete(item)
        db.commit()
    return {"message": "CPC eliminado del catálogo"}


@app.delete("/pac/cpc/catalog-all")
def pac_cpc_delete_all(user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    db.query(CPCCatalog).delete()
    db.commit()
    return {"message": "Catálogo CPC vaciado"}


# CPC Loaded Data endpoints (persistencia temporal de datos cargados desde Excel)

@app.get("/pac/cpc/loaded")
def pac_cpc_loaded_get(user: User = Depends(require_module("pac")), db: Session = Depends(get_db)):
    items = db.query(CPCLoadedData).order_by(CPCLoadedData.id).all()
    return [CPCLoadedDataResponse.model_validate(i).model_dump() for i in items]


@app.post("/pac/cpc/loaded")
def pac_cpc_loaded_save(data: dict, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    items = data.get("items", [])
    if not items or not isinstance(items, list):
        raise HTTPException(status_code=400, detail="Se requiere una lista de items")

    db.query(CPCLoadedData).delete()
    for item in items:
        cpc_code = str(item.get("cpc", "")).strip()
        if not cpc_code:
            continue
        new_item = CPCLoadedData(
            cpc=cpc_code,
            descripcion=item.get("descripcion", ""),
            umbral=float(item.get("umbral", 0) or 0),
        )
        db.add(new_item)
    db.commit()
    return {"message": f"{len(items)} registros cargados guardados"}


@app.delete("/pac/cpc/loaded/{cpc_code}")
def pac_cpc_loaded_delete(cpc_code: str, user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    item = db.query(CPCLoadedData).filter(CPCLoadedData.cpc == cpc_code).first()
    if item:
        db.delete(item)
        db.commit()
    return {"message": "Registro cargado eliminado"}


@app.delete("/pac/cpc/loaded")
def pac_cpc_loaded_clear(user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    db.query(CPCLoadedData).delete()
    db.commit()
    return {"message": "Datos cargados eliminados"}


# ==================== CAM (Cambio de Administrador) Endpoints ====================

@app.post("/cam/extract")
async def cam_upload_file(file: UploadFile = File(...), user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    if not file:
        raise HTTPException(status_code=400, detail="No se envió archivo")
    filename = file.filename or "unknown"
    ext = filename.lower().split(".")[-1] if "." in filename else ""
    if ext != "pdf":
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos PDF")

    upload_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    filepath = os.path.join(upload_dir, f"cam_{uuid.uuid4().hex}_{filename}")
    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    try:
        from services.cam_service import extract_cam_data
        data = extract_cam_data(filepath)
        if data.get("codigo_proceso"):
            result = process_cam_pdf(filepath, filename, db)
        else:
            result = {
                "id": None,
                "filename": filename,
                "codigo_proceso": data.get("codigo_proceso"),
                "administrador_contrato_actual": data.get("administrador_contrato_actual"),
                "objeto_proceso": data.get("objeto_proceso"),
                "estado_proceso": data.get("estado_proceso"),
                "fecha_procesamiento": datetime.utcnow().isoformat(),
            }
        log_audit(user.id, "CAM_EXTRACT", "cam_extractions", details=f"Archivo: {filename}", db=db)
        return result
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Error al procesar PDF CAM: {str(e)}")


@app.get("/cam/extractions")
def cam_list(user: User = Depends(require_module("cam")), db: Session = Depends(get_db)):
    return cam_list_extractions(db)


@app.get("/procesos/administradores")
def procesos_administradores(search: Optional[str] = Query(None), user: User = Depends(require_auth), db: Session = Depends(get_db)):
    query = db.query(CAMExtraction).filter(CAMExtraction.administrador_contrato_actual.isnot(None)).filter(CAMExtraction.administrador_contrato_actual != "")
    if search:
        query = query.filter(CAMExtraction.administrador_contrato_actual.ilike(f"%{search}%"))
    extractions = query.order_by(CAMExtraction.administrador_contrato_actual).all()
    grouped: dict = {}
    for ext in extractions:
        admin = ext.administrador_contrato_actual.strip()
        if admin not in grouped:
            grouped[admin] = {
                "administrador": admin,
                "procesos": [],
                "total_procesos": 0,
            }
        grouped[admin]["procesos"].append({
            "id": ext.id,
            "codigo_proceso": ext.codigo_proceso or "",
            "objeto_proceso": ext.objeto_proceso or "",
            "estado_proceso": ext.estado_proceso or "",
            "filename": ext.filename or "",
            "fecha_publicacion": ext.fecha_publicacion or "",
            "fecha_procesamiento": ext.fecha_procesamiento or "",
        })
        grouped[admin]["total_procesos"] += 1
    return list(grouped.values())

@app.get("/cam/extractions/{extraction_id}")
def cam_get(extraction_id: str, user: User = Depends(require_module("cam")), db: Session = Depends(get_db)):
    ext = cam_get_extraction(extraction_id, db)
    if not ext:
        raise HTTPException(status_code=404, detail="Extracción no encontrada")
    return ext


@app.put("/cam/extractions/{extraction_id}")
def cam_update(extraction_id: str, data: dict, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    from services.cam_service import update_extraction
    try:
        result = update_extraction(extraction_id, data, db)
        if not result:
            raise HTTPException(status_code=404, detail="Extracción no encontrada")
        log_audit(user.id, "UPDATE", "cam_extractions", resource_id=extraction_id, db=db)
        return result
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))

@app.delete("/cam/extractions/{extraction_id}")
def cam_delete(extraction_id: str, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    if not cam_delete_extraction(extraction_id, db):
        raise HTTPException(status_code=404, detail="Extracción no encontrada")
    log_audit(user.id, "CAM_DELETE", "cam_extractions", resource_id=extraction_id, db=db)
    return {"status": "ok", "message": "Extracción eliminada"}


# ==================== Catálogo Electrónico Endpoints ====================

@app.post("/ce/upload")
async def ce_upload_file(file: UploadFile = File(...), user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    if not file:
        raise HTTPException(status_code=400, detail="No se envió archivo")
    filename = file.filename or "unknown"
    ext = filename.lower().split(".")[-1] if "." in filename else ""
    if ext != "pdf":
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos PDF")

    upload_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    filepath = os.path.join(upload_dir, f"ce_{uuid.uuid4().hex}_{filename}")
    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    try:
        result = process_pdf(filepath, filename, db)
        log_audit(user.id, "UPLOAD", "ce_extractions", details=f"Archivo: {filename}", db=db)
        return result
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Error al procesar PDF: {str(e)}")


@app.get("/ce/extractions")
def ce_list_extractions(user: User = Depends(require_module("ce")), db: Session = Depends(get_db)):
    return list_extractions(db)


@app.get("/ce/extractions/{extraction_id}")
def ce_get_extraction(extraction_id: str, user: User = Depends(require_module("ce")), db: Session = Depends(get_db)):
    ext = get_extraction(extraction_id, db)
    if not ext:
        raise HTTPException(status_code=404, detail="Extracción no encontrada")
    return ext


@app.put("/ce/extractions/{extraction_id}")
def ce_update_extraction(extraction_id: str, data: dict, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    from services.ce_service import update_extraction
    result = update_extraction(extraction_id, data, db)
    if not result:
        raise HTTPException(status_code=404, detail="Extracción no encontrada")
    log_audit(user.id, "UPDATE", "ce_extractions", resource_id=extraction_id, db=db)
    return result

@app.delete("/ce/extractions/{extraction_id}")
def ce_delete_extraction(extraction_id: str, user: User = Depends(require_role("admin", "operator")), db: Session = Depends(get_db)):
    if not delete_extraction(extraction_id, db):
        raise HTTPException(status_code=404, detail="Extracción no encontrada")
    log_audit(user.id, "DELETE", "ce_extractions", resource_id=extraction_id, db=db)
    return {"status": "ok", "message": "Extracción eliminada"}


@app.delete("/ce/extractions")
def ce_clear_extractions(user: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    clear_all(db)
    log_audit(user.id, "DELETE_ALL", "ce_extractions", db=db)
    return {"status": "ok", "message": "Todas las extracciones eliminadas"}


@app.get("/ce/export-excel")
def ce_export_excel(ids: Optional[str] = Query(None, description="Comma-separated extraction IDs"), user: User = Depends(require_auth), db: Session = Depends(get_db)):
    try:
        extraction_ids = ids.split(",") if ids else None
        filepath = export_excel(extraction_ids, db)
        filename = os.path.basename(filepath)
        return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar: {str(e)}")


@app.get("/ce/export-excel-by-admin/{admin_name}")
def ce_export_excel_by_admin(admin_name: str, user: User = Depends(require_auth), db: Session = Depends(get_db)):
    try:
        from urllib.parse import unquote
        admin_name = unquote(admin_name)
        filepath = export_excel_by_admin(admin_name, db)
        filename = os.path.basename(filepath)
        return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar: {str(e)}")


@app.get("/procesos/export-excel-by-admin")
def procesos_export_excel_by_admin(admin_name: str = Query(...), user: User = Depends(require_auth), db: Session = Depends(get_db)):
    from services.cam_service import export_excel_by_admin
    try:
        filepath = export_excel_by_admin(admin_name, db)
        filename = os.path.basename(filepath)
        return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar: {str(e)}")


# ==================== Frontend SPA (single-deploy) ====================
# Serve the built React app when frontend/dist exists. Must be declared
# after all API routes so that /api routes win and SPA deep links fall
# back to index.html.
from fastapi.responses import FileResponse as _FileResponse
from fastapi.staticfiles import StaticFiles as _StaticFiles

_SPA_DIST = os.path.abspath(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "frontend", "dist"))
if os.path.isdir(_SPA_DIST):
    _spa_assets = os.path.join(_SPA_DIST, "assets")
    if os.path.isdir(_spa_assets):
        app.mount("/assets", _StaticFiles(directory=_spa_assets), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    def _spa_fallback(full_path: str):
        candidate = os.path.join(_SPA_DIST, full_path)
        if full_path and os.path.isfile(candidate):
            return _FileResponse(candidate)
        return _FileResponse(os.path.join(_SPA_DIST, "index.html"))


if __name__ == "__main__":
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8000
    uvicorn.run("main:app", host="127.0.0.1", port=port, reload=False)
